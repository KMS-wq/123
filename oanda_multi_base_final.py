#!/usr/bin/env python3
"""
Oanda Currency Rate Scraper - Улучшенная версия
С автоматическим принятием cookies и переключением в табличный вид
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import time
import json
import csv
import re
from datetime import datetime
import sys


class OandaRateScraperV2:
    """Улучшенная версия скрапера для Oanda"""
    
    def __init__(self, headless=False, wait_time=20):
        """
        Инициализация
        
        Args:
            headless (bool): Режим без GUI
            wait_time (int): Время ожидания загрузки элементов (секунды)
        """
        self.url = "https://fxds-hcc.oanda.com/"
        self.wait_time = wait_time
        self.driver = None
        
        # Настройка Chrome
        self.options = Options()
        if headless:
            self.options.add_argument('--headless=new')
        
        self.options.add_argument('--no-sandbox')
        self.options.add_argument('--disable-dev-shm-usage')
        self.options.add_argument('--disable-gpu')
        self.options.add_argument('--disable-blink-features=AutomationControlled')
        self.options.add_argument('--window-size=1920,1080')
        self.options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        # Отключаем уведомления
        self.options.add_experimental_option("excludeSwitches", ["enable-automation"])
        self.options.add_experimental_option('useAutomationExtension', False)
        self.options.add_experimental_option("prefs", {
            "profile.default_content_setting_values.notifications": 2
        })
    
    def start(self):
        """Запуск браузера"""
        print("🚀 Запуск браузера Chrome...")
        try:
            self.driver = webdriver.Chrome(options=self.options)
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            print("✅ Браузер запущен успешно")
            return True
        except WebDriverException as e:
            print(f"❌ Ошибка запуска браузера: {e}")
            return False
    
    def load_page(self):
        """Загрузка страницы Oanda"""
        print(f"📄 Загрузка страницы: {self.url}")
        try:
            self.driver.get(self.url)
            time.sleep(3)
            print("✅ Страница загружена")
            return True
        except Exception as e:
            print(f"❌ Ошибка загрузки страницы: {e}")
            return False
    
    def accept_cookies(self):
        """Принять cookies"""
        print("🍪 Принимаю cookies...")
        try:
            wait = WebDriverWait(self.driver, 10)
            
            # Возможные селекторы для кнопки Accept Cookies
            cookie_selectors = [
                "//button[contains(text(), 'Accept All Cookies')]",
                "//button[contains(text(), 'Accept')]",
                "//button[@id='onetrust-accept-btn-handler']",
                "button[class*='accept']",
                "#onetrust-accept-btn-handler"
            ]
            
            for selector in cookie_selectors:
                try:
                    if selector.startswith("//"):
                        button = wait.until(EC.element_to_be_clickable((By.XPATH, selector)))
                    else:
                        button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector)))
                    
                    button.click()
                    print("✅ Cookies приняты")
                    time.sleep(2)
                    return True
                except:
                    continue
            
            print("⚠️ Кнопка cookies не найдена (возможно уже принято)")
            return True
            
        except Exception as e:
            print(f"⚠️ Не удалось принять cookies: {e}")
            return True  # Продолжаем в любом случае
    
    def select_time_period(self, days=180):
        """
        Выбор периода времени для отображения данных
        
        Args:
            days (int): Количество дней (30, 90, 180, 365 и т.д.)
        """
        print(f"📅 Выбираю период: последние {days} дней...")
        
        # Маппинг периодов на индексы в списке (li[X])
        period_indices = {
            30: 1,    # li[1] - 30 дней
            60: 2,    # li[2] - 60 дней
            90: 3,    # li[3] - 90 дней
            120: 4,   # li[4] - 120 дней
            180: 5,   # li[5] - 180 дней
            365: 6,   # li[6] - 1 год
            730: 7,   # li[7] - 2 года
        }
        
        period_index = period_indices.get(days, 5)  # По умолчанию 180 дней
        
        try:
            wait = WebDriverWait(self.driver, 10)
            
            # Шаг 1: Клик на выпадающий список периодов
            print("  1️⃣ Открываю выбор периода...")
            step1_path = '/html/body/div[2]/div[2]/div[1]/div[1]/ul/li[1]/div/div[1]/div/div'
            try:
                element1 = wait.until(EC.element_to_be_clickable((By.XPATH, step1_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", element1)
                time.sleep(0.5)
                element1.click()
                print("  ✅ Выпадающий список открыт")
                time.sleep(1)
            except Exception as e:
                print(f"  ⚠️ Не удалось открыть список: {e}")
                return False
            
            # Шаг 2: Клик на поле ввода (активация)
            print("  2️⃣ Активирую поле...")
            step2_path = '/html/body/div[2]/div[2]/div[1]/div[1]/ul/li[1]/div/div[2]/div[1]/div[1]/div[1]/input'
            try:
                element2 = wait.until(EC.element_to_be_clickable((By.XPATH, step2_path)))
                element2.click()
                time.sleep(1)
                print("  ✅ Поле активировано")
            except Exception as e:
                print(f"  ⚠️ Не удалось активировать: {e}")
            
            # Шаг 3: Выбор периода из списка
            print(f"  3️⃣ Выбираю {days} дней (элемент {period_index})...")
            step3_path = f'/html/body/div[2]/div[2]/div[1]/div[1]/ul/li[1]/div/div[2]/div[1]/div[1]/div[2]/ul/li[{period_index}]/div/div'
            try:
                element3 = wait.until(EC.element_to_be_clickable((By.XPATH, step3_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", element3)
                time.sleep(0.5)
                element3.click()
                print(f"  ✅ Период {days} дней выбран")
                time.sleep(1)
            except Exception as e:
                print(f"  ⚠️ Не удалось выбрать период: {e}")
                # Пробуем альтернативный метод - поиск по тексту
                try:
                    print(f"  🔄 Пробую найти период по тексту...")
                    period_element = self.driver.find_element(By.XPATH, f"//div[contains(text(), '{days}')]")
                    period_element.click()
                    print(f"  ✅ Период найден и выбран")
                    time.sleep(1)
                except:
                    print(f"  ❌ Не удалось найти период {days} дней")
                    return False
            
            # Шаг 4: Подтверждение выбора (кнопка Apply/OK)
            print("  4️⃣ Подтверждаю выбор...")
            step4_path = '/html/body/div[2]/div[2]/div[1]/div[1]/ul/li[1]/div/div[2]/div[3]/button[1]'
            try:
                element4 = wait.until(EC.element_to_be_clickable((By.XPATH, step4_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", element4)
                time.sleep(0.5)
                element4.click()
                print("  ✅ Выбор подтвержден")
                time.sleep(3)  # Ждем обновления данных
            except Exception as e:
                print(f"  ⚠️ Не удалось подтвердить: {e}")
                return False
            
            print(f"✅ Период {days} дней успешно выбран")
            return True
            
        except Exception as e:
            print(f"❌ Ошибка при выборе периода: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def select_currency_pair(self, base_currency="USD", target_currency="EUR"):
        """
        Выбор валютной пары
        
        Args:
            base_currency (str): Базовая валюта (по умолчанию USD)
            target_currency (str): Целевая валюта (по умолчанию EUR)
        """
        print(f"💱 Выбираю валютную пару: {base_currency}/{target_currency}...")
        
        # Словарь валют и их индексов в списке
        currency_indices = {
            "AED": 10,   # li[10] - UAE Dirham
            "AUD": 15,   # Австралийский доллар
            "EUR": 5,    # Евро
            "GBP": 8,    # Британский фунт
            "JPY": 12,   # Японская иена
            "CHF": 7,    # Швейцарский франк
            "CNY": 20,   # Китайский юань
            "CAD": 18,   # Канадский доллар
            "RUB": 25,   # Российский рубль
            # Добавьте другие валюты по необходимости
        }
        
        target_index = currency_indices.get(target_currency, 10)
        
        try:
            wait = WebDriverWait(self.driver, 10)
            
            # Шаг 1: Клик на поле ввода второй валюты
            print(f"  1️⃣ Открываю выбор второй валюты ({target_currency})...")
            input_path = '/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div[2]/div[2]/div[2]/div/input'
            try:
                input_element = wait.until(EC.element_to_be_clickable((By.XPATH, input_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", input_element)
                time.sleep(0.5)
                input_element.click()
                print("  ✅ Поле ввода активировано")
                time.sleep(1)
            except Exception as e:
                print(f"  ⚠️ Не удалось активировать поле: {e}")
                return False
            
            # Шаг 2: Выбор валюты из списка
            print(f"  2️⃣ Выбираю {target_currency} (элемент {target_index})...")
            currency_path = f'/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div[2]/div[2]/div[2]/div/ul/li[{target_index}]/div[1]'
            try:
                currency_element = wait.until(EC.element_to_be_clickable((By.XPATH, currency_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", currency_element)
                time.sleep(0.5)
                currency_element.click()
                print(f"  ✅ {target_currency} выбран")
                time.sleep(2)  # Ждем загрузки данных
            except Exception as e:
                print(f"  ⚠️ Не удалось выбрать валюту: {e}")
                # Пробуем найти по тексту
                try:
                    print(f"  🔄 Пробую найти {target_currency} по тексту...")
                    currency_text = self.driver.find_element(By.XPATH, f"//div[contains(text(), '{target_currency}')]")
                    currency_text.click()
                    print(f"  ✅ {target_currency} найден и выбран")
                    time.sleep(2)
                except:
                    print(f"  ❌ Не удалось найти {target_currency}")
                    return False
            
            print(f"✅ Валютная пара {base_currency}/{target_currency} успешно выбрана")
            return True
            
        except Exception as e:
            print(f"❌ Ошибка при выборе валютной пары: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def add_currency_column(self, column_number, currency_code, li_index):
        """
        Добавить валюту в определенный столбец
        
        Args:
            column_number (int): Номер столбца (3, 4, 5, 6...)
            currency_code (str): Код валюты (CHF, GBP, JPY и т.д.)
            li_index (int): Индекс элемента li в списке
        """
        print(f"➕ Добавляю {currency_code} в столбец {column_number}...")
        
        try:
            wait = WebDriverWait(self.driver, 10)
            
            # Шаг 1: Клик на контейнер валюты (открывает список)
            container_path = f'/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div[2]/div[{column_number}]'
            print(f"  1️⃣ Открываю список столбца {column_number}...")
            
            try:
                container = wait.until(EC.element_to_be_clickable((By.XPATH, container_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", container)
                time.sleep(0.5)
                container.click()
                print(f"  ✅ Список открыт")
                time.sleep(1)
            except Exception as e:
                print(f"  ⚠️ Не удалось открыть список: {e}")
                return False
            
            # Шаг 2: Выбор валюты из списка
            currency_path = f'/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div[2]/div[{column_number}]/div[2]/div/ul/li[{li_index}]/div[1]'
            print(f"  2️⃣ Выбираю {currency_code} (li[{li_index}])...")
            
            try:
                currency_element = wait.until(EC.element_to_be_clickable((By.XPATH, currency_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", currency_element)
                time.sleep(0.5)
                currency_element.click()
                print(f"  ✅ {currency_code} выбран")
                time.sleep(3)  # Ждем обновления данных
                return True
            except Exception as e:
                print(f"  ⚠️ Не удалось выбрать {currency_code}: {e}")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка добавления {currency_code}: {e}")
            return False
    
    def change_base_currency(self, currency_code, li_index):
        """
        Сменить базовую валюту (первый столбец)
        
        Args:
            currency_code (str): Код валюты (EUR, AED, GBP, HKD)
            li_index (int): Индекс в списке
        """
        print(f"🔄 Меняю базовую валюту на {currency_code}...")
        
        try:
            wait = WebDriverWait(self.driver, 10)
            
            # Шаг 1: Клик на контейнер базовой валюты
            base_currency_path = '/html/body/div[2]/div[2]/div[1]/div[2]/div[1]/div[2]'
            print(f"  1️⃣ Открываю список базовых валют...")
            
            try:
                base_container = wait.until(EC.element_to_be_clickable((By.XPATH, base_currency_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", base_container)
                time.sleep(0.5)
                base_container.click()
                print(f"  ✅ Список открыт")
                time.sleep(1)
            except Exception as e:
                print(f"  ⚠️ Не удалось открыть список: {e}")
                return False
            
            # Шаг 2: Выбор валюты из списка
            currency_path = f'/html/body/div[2]/div[2]/div[1]/div[2]/div[1]/div[2]/div/div/ul/li[{li_index}]/div[1]'
            print(f"  2️⃣ Выбираю {currency_code} (li[{li_index}])...")
            
            try:
                currency_element = wait.until(EC.element_to_be_clickable((By.XPATH, currency_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", currency_element)
                time.sleep(0.5)
                currency_element.click()
                print(f"  ✅ Базовая валюта изменена на {currency_code}")
                time.sleep(4)  # Ждем обновления всех данных
                return True
            except Exception as e:
                print(f"  ⚠️ Не удалось выбрать {currency_code}: {e}")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка смены базовой валюты: {e}")
            return False
    
    def select_currency_in_new_column(self, div_number, currency_code, li_index):
        """
        Выбрать валюту в новом столбце (после нажатия '+')
        
        Args:
            div_number (int): Номер div столбца (6, 7, 8, 9, 10)
            currency_code (str): Код валюты (INR, RUB, KZT, HKD, USD)
            li_index (int): Индекс li в списке
        """
        print(f"💱 Выбираю {currency_code} в столбце div[{div_number}]...")
        
        try:
            wait = WebDriverWait(self.driver, 10)
            
            # Шаг 1: Клик на контейнер столбца (открывает список)
            container_path = f'/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div[2]/div[{div_number}]'
            print(f"  1️⃣ Открываю список столбца {div_number}...")
            
            try:
                container = wait.until(EC.element_to_be_clickable((By.XPATH, container_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", container)
                time.sleep(0.5)
                container.click()
                print(f"  ✅ Список открыт")
                time.sleep(1)
            except Exception as e:
                print(f"  ⚠️ Не удалось открыть список: {e}")
                return False
            
            # Шаг 2: Выбор валюты из списка
            currency_path = f'/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div[2]/div[{div_number}]/div[2]/div/ul/li[{li_index}]/div[1]'
            print(f"  2️⃣ Выбираю {currency_code} (li[{li_index}])...")
            
            try:
                currency_element = wait.until(EC.element_to_be_clickable((By.XPATH, currency_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", currency_element)
                time.sleep(0.5)
                currency_element.click()
                print(f"  ✅ {currency_code} выбран")
                time.sleep(3)
                return True
            except Exception as e:
                print(f"  ⚠️ Не удалось выбрать {currency_code}: {e}")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка выбора валюты {currency_code}: {e}")
            return False
    
    def add_multiple_plus_buttons(self, plus_div_numbers):
        """
        Нажать несколько кнопок '+' подряд для добавления новых столбцов
        
        Args:
            plus_div_numbers (list): Список номеров div для кнопок '+' [6, 7, 8, 9, 10]
        """
        print(f"➕ Добавляю {len(plus_div_numbers)} новых столбцов через '+'...")
        
        try:
            wait = WebDriverWait(self.driver, 10)
            
            for idx, div_num in enumerate(plus_div_numbers, 1):
                print(f"\n  {idx}/{len(plus_div_numbers)} - Нажимаю '+' на div[{div_num}]...")
                
                plus_button_path = f'/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div[2]/div[{div_num}]'
                
                try:
                    plus_button = wait.until(EC.element_to_be_clickable((By.XPATH, plus_button_path)))
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", plus_button)
                    time.sleep(0.5)
                    plus_button.click()
                    print(f"  ✅ Столбец {idx} добавлен")
                    time.sleep(1.5)
                except Exception as e:
                    print(f"  ⚠️ Не удалось нажать '+' на div[{div_num}]: {e}")
                    return False
            
            print(f"\n✅ Все {len(plus_div_numbers)} столбцов добавлены!")
            return True
            
        except Exception as e:
            print(f"❌ Ошибка при добавлении столбцов: {e}")
            return False
        """
        Добавить дополнительную валюту (нужно сначала нажать кнопку '+')
        
        Args:
            currency_code (str): Код валюты
            li_index (int): Индекс элемента li в списке
            plus_button_div_number (int): Номер div для кнопки '+' (6, 7, 8, 9...)
        """
        print(f"➕ Добавляю дополнительную валюту {currency_code}...")
        
        try:
            wait = WebDriverWait(self.driver, 10)
            
            # Шаг 1: Клик на кнопку '+' для добавления нового столбца
            plus_button_path = f'/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div[2]/div[{plus_button_div_number}]'
            print(f"  1️⃣ Нажимаю кнопку '+' (div[{plus_button_div_number}])...")
            
            try:
                plus_button = wait.until(EC.element_to_be_clickable((By.XPATH, plus_button_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", plus_button)
                time.sleep(0.5)
                plus_button.click()
                print(f"  ✅ Новый столбец добавлен")
                time.sleep(1)
            except Exception as e:
                print(f"  ⚠️ Не удалось нажать '+': {e}")
                return False
            
            # Шаг 2: Выбор валюты из списка нового столбца
            # После нажатия '+' новый столбец появляется с тем же div номером
            currency_path = f'/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div[2]/div[{plus_button_div_number}]/div[2]/div/ul/li[{li_index}]/div[1]'
            print(f"  2️⃣ Выбираю {currency_code} (li[{li_index}])...")
            
            try:
                currency_element = wait.until(EC.element_to_be_clickable((By.XPATH, currency_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", currency_element)
                time.sleep(0.5)
                currency_element.click()
                print(f"  ✅ {currency_code} выбран")
                time.sleep(3)
                return True
            except Exception as e:
                print(f"  ⚠️ Не удалось выбрать {currency_code}: {e}")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка добавления дополнительной валюты {currency_code}: {e}")
            return False
        """
        Выбор валютной пары (например USD/EUR, USD/AED и т.д.)
        
        Args:
            base_currency (str): Базовая валюта (USD, EUR и т.д.)
            quote_currency (str): Котируемая валюта (EUR, AED, JPY и т.д.)
        """
        print(f"💱 Выбираю валютную пару {base_currency}/{quote_currency}...")
        
        # Словарь индексов валют в списке
        currency_list_indices = {
            "AED": 10,   # li[10] - UAE Dirham
            "EUR": 1,    # li[1] - Euro (примерно)
            "GBP": 2,    # li[2] - British Pound
            "JPY": 3,    # li[3] - Japanese Yen
            "CHF": 4,    # li[4] - Swiss Franc
            "AUD": 5,    # li[5] - Australian Dollar
            "CAD": 6,    # li[6] - Canadian Dollar
            "CNY": 7,    # li[7] - Chinese Yuan
        }
        
        try:
            wait = WebDriverWait(self.driver, 10)
            
            # Шаг 1: Клик на контейнер валюты (открывает выпадающий список)
            print(f"  1️⃣ Открываю список валют...")
            container_path = '/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div[2]/div[2]'
            try:
                container = wait.until(EC.element_to_be_clickable((By.XPATH, container_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", container)
                time.sleep(0.5)
                container.click()
                print("  ✅ Список валют открыт")
                time.sleep(1)
            except Exception as e:
                print(f"  ⚠️ Не удалось открыть список: {e}")
                return False
            
            # Шаг 2: Выбор валюты из списка
            quote_index = currency_list_indices.get(quote_currency, 10)
            print(f"  2️⃣ Выбираю {quote_currency} (элемент {quote_index})...")
            
            currency_path = f'/html/body/div[2]/div[2]/div[1]/div[2]/div[2]/div[2]/div[2]/div[2]/div/ul/li[{quote_index}]/div[1]'
            try:
                currency_element = wait.until(EC.element_to_be_clickable((By.XPATH, currency_path)))
                self.driver.execute_script("arguments[0].scrollIntoView(true);", currency_element)
                time.sleep(0.5)
                currency_element.click()
                print(f"  ✅ Валюта {quote_currency} выбрана")
                time.sleep(4)  # Увеличенное время ожидания обновления данных
            except Exception as e:
                print(f"  ⚠️ Не удалось выбрать валюту: {e}")
                
                # Пробуем альтернативный метод - поиск по тексту
                try:
                    print(f"  🔄 Ищу {quote_currency} по тексту...")
                    currency_elem = self.driver.find_element(By.XPATH, f"//div[contains(text(), '{quote_currency}')]")
                    currency_elem.click()
                    print(f"  ✅ {quote_currency} найден и выбран")
                    time.sleep(2)
                except:
                    print(f"  ❌ Не удалось найти {quote_currency}")
                    return False
            
            print(f"✅ Валютная пара {base_currency}/{quote_currency} успешно выбрана")
            return True
            
        except Exception as e:
            print(f"❌ Ошибка при выборе валютной пары: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def switch_to_table_view(self):
        """Переключиться в табличный вид"""
        print("📊 Переключаюсь в табличный вид...")
        try:
            wait = WebDriverWait(self.driver, 15)
            
            # ПРАВИЛЬНЫЕ пути к кнопке таблицы
            table_selectors = [
                '/html/body/div[2]/div[2]/div[1]/div[3]/div/div[1]/div[1]/div[2]',  # Full XPath
                '//*[@id="hcc"]/div[2]/div[1]/div[3]/div/div[1]/div[1]/div[2]',     # Относительный XPath
                "//div[contains(@class, 'interbank')]//div[2]",
                "//button[contains(@class, 'table')]",
                "//button[contains(@aria-label, 'table')]",
                "button[class*='table']"
            ]
            
            for selector in table_selectors:
                try:
                    print(f"  🔍 Пробую селектор: {selector[:50]}...")
                    
                    if selector.startswith("//") or selector.startswith('/html'):
                        element = wait.until(EC.presence_of_element_located((By.XPATH, selector)))
                    else:
                        element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                    
                    # Прокручиваем к элементу
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", element)
                    time.sleep(0.5)
                    
                    # Пробуем обычный клик
                    try:
                        element.click()
                        print("✅ Переключено в табличный вид (обычный клик)")
                        time.sleep(3)
                        return True
                    except:
                        # Если не получилось, пробуем JavaScript клик
                        self.driver.execute_script("arguments[0].click();", element)
                        print("✅ Переключено в табличный вид (JavaScript клик)")
                        time.sleep(3)
                        return True
                except Exception as e:
                    print(f"  ⚠️ Не подошел: {str(e)[:50]}")
                    continue
            
            # Пробуем найти все кнопки и кликнуть на вторую
            print("  🔍 Пробую найти кнопки через поиск всех элементов...")
            try:
                buttons = self.driver.find_elements(By.TAG_NAME, "button")
                print(f"  📋 Найдено кнопок на странице: {len(buttons)}")
                
                for i, btn in enumerate(buttons):
                    btn_class = btn.get_attribute('class') or ''
                    btn_aria = btn.get_attribute('aria-label') or ''
                    
                    # Ищем кнопку с иконкой таблицы
                    if 'table' in btn_class.lower() or 'table' in btn_aria.lower():
                        print(f"  ✅ Найдена кнопка таблицы (индекс {i})")
                        self.driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                        time.sleep(0.5)
                        self.driver.execute_script("arguments[0].click();", btn)
                        print("✅ Переключено в табличный вид")
                        time.sleep(3)
                        return True
            except Exception as e:
                print(f"  ⚠️ Ошибка поиска кнопок: {e}")
            
            print("⚠️ Кнопка переключения не найдена, возможно уже в табличном виде")
            print("💡 Делаю скриншот для проверки...")
            self.driver.save_screenshot("debug_table_button.png")
            return True
            
        except Exception as e:
            print(f"⚠️ Не удалось переключить вид: {e}")
            return True
    
    def extract_rates_from_table(self):
        """Извлечение курсов из таблицы"""
        print("💰 Извлекаю курсы валют из таблицы...")
        rates = {}
        
        try:
            wait = WebDriverWait(self.driver, 10)
            
            # ТОЧНЫЙ путь к tbody с данными
            tbody_selectors = [
                '/html/body/div[2]/div[2]/div[3]/div[2]/table/tbody',  # Full XPath
                '//div[3]/div[2]/table/tbody',  # Короче
                'table tbody',  # Общий
            ]
            
            tbody = None
            for selector in tbody_selectors:
                try:
                    print(f"  🔍 Пробую селектор tbody: {selector}")
                    if selector.startswith("//") or selector.startswith('/html'):
                        tbody = wait.until(EC.presence_of_element_located((By.XPATH, selector)))
                    else:
                        tbody = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                    
                    if tbody:
                        print(f"  ✅ Таблица найдена!")
                        break
                except:
                    continue
            
            if not tbody:
                print("  ⚠️ tbody не найден, ищу таблицы...")
                tables = self.driver.find_elements(By.TAG_NAME, "table")
                print(f"  📊 Найдено таблиц: {len(tables)}")
                
                for i, table in enumerate(tables):
                    try:
                        tbody = table.find_element(By.TAG_NAME, "tbody")
                        print(f"  ✅ Использую tbody из таблицы {i+1}")
                        break
                    except:
                        continue
            
            if tbody:
                # Извлекаем все строки из tbody
                rows = tbody.find_elements(By.TAG_NAME, "tr")
                print(f"  📋 Найдено строк данных: {len(rows)}")
                
                for i, row in enumerate(rows):
                    try:
                        cells = row.find_elements(By.TAG_NAME, "td")
                        
                        if len(cells) >= 2:
                            # Извлекаем текст из всех ячеек
                            cell_data = [cell.text.strip() for cell in cells]
                            
                            # Первая ячейка может быть валютная пара
                            currency_pair = cell_data[0]
                            
                            # Ищем курс в ячейках (обычно во второй или третьей)
                            for idx, cell_text in enumerate(cell_data[1:], 1):
                                if cell_text and self._is_rate(cell_text):
                                    # Проверяем, что это валютная пара
                                    if self._is_currency_pair(currency_pair):
                                        rates[currency_pair] = cell_text
                                        print(f"    ✅ {currency_pair}: {cell_text}")
                                        break
                                    # Или может быть дата + курс
                                    elif '/' in currency_pair or '-' in currency_pair:
                                        # Возможно это дата, берем следующую ячейку как пару
                                        if idx < len(cell_data) - 1:
                                            next_cell = cell_data[idx + 1]
                                            if self._is_currency_pair(next_cell):
                                                rates[next_cell] = cell_text
                                                print(f"    ✅ {next_cell}: {cell_text}")
                                                break
                            
                            # Если ничего не нашли, выводим содержимое строки для отладки
                            if i < 5:  # Первые 5 строк для отладки
                                print(f"    📄 Строка {i+1}: {cell_data}")
                    
                    except Exception as e:
                        print(f"    ⚠️ Ошибка в строке {i+1}: {e}")
                        continue
            
            if rates:
                print(f"\n  ✅ Успешно извлечено курсов: {len(rates)}")
            else:
                print(f"\n  ⚠️ Курсы не извлечены, делаю скриншот таблицы...")
                self.driver.save_screenshot("debug_table_content.png")
            
            return rates
            
        except Exception as e:
            print(f"❌ Ошибка при извлечении из таблицы: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    def extract_rates_alternative(self):
        """Альтернативный метод извлечения курсов"""
        print("🔍 Использую альтернативный метод поиска...")
        rates = {}
        
        try:
            # Метод 1: Поиск в HTML по паттернам
            html = self.driver.page_source
            
            patterns = [
                r'([A-Z]{3}/[A-Z]{3})["\s<>:]+([0-9]+\.[0-9]{4,})',
                r'currency["\s:]+([A-Z]{3}/[A-Z]{3})["\s,]+rate["\s:]+([0-9]+\.[0-9]+)',
            ]
            
            for pattern in patterns:
                matches = re.findall(pattern, html)
                for pair, rate in matches:
                    if self._is_currency_pair(pair):
                        rates[pair] = rate
            
            # Метод 2: Поиск всех элементов с текстом валютных пар
            elements = self.driver.find_elements(By.XPATH, "//*[contains(text(), 'USD') or contains(text(), 'EUR')]")
            
            for elem in elements:
                text = elem.text.strip()
                match = re.search(r'([A-Z]{3}/[A-Z]{3})\s*([0-9]+\.[0-9]+)', text)
                if match:
                    rates[match.group(1)] = match.group(2)
            
            return rates
            
        except Exception as e:
            print(f"⚠️ Ошибка альтернативного метода: {e}")
            return {}
    
    def _is_currency_pair(self, text):
        """Проверка валютной пары"""
        return bool(re.match(r'^[A-Z]{3}/[A-Z]{3}$', text))
    
    def extract_full_table_data(self):
        """Извлечение всех данных из таблицы включая даты"""
        print("📊 Извлекаю полные данные таблицы...")
        table_data = []
        
        try:
            # Точный путь к tbody
            tbody_path = '/html/body/div[2]/div[2]/div[3]/div[2]/table/tbody'
            
            try:
                tbody = self.driver.find_element(By.XPATH, tbody_path)
            except:
                # Запасной вариант
                tbody = self.driver.find_element(By.CSS_SELECTOR, 'table tbody')
            
            rows = tbody.find_elements(By.TAG_NAME, "tr")
            print(f"  📋 Найдено строк: {len(rows)}")
            
            for i, row in enumerate(rows):
                cells = row.find_elements(By.TAG_NAME, "td")
                row_data = [cell.text.strip() for cell in cells]
                
                if row_data:
                    table_data.append(row_data)
                    if i < 5:  # Выводим первые 5 строк
                        print(f"    Строка {i+1}: {row_data}")
            
            return table_data
            
        except Exception as e:
            print(f"❌ Ошибка извлечения полных данных: {e}")
            return []
    
    def parse_table_data(self, raw_table_data, num_currencies):
        """
        Парсинг сырых данных таблицы и разделение по валютам
        
        Args:
            raw_table_data (list): Сырые данные из таблицы
            num_currencies (int): Количество валют (столбцов)
            
        Returns:
            list: Структурированные данные [[дата, курс1, курс2, ...], ...]
        """
        print(f"🔍 Парсинг данных для {num_currencies} валют...")
        
        parsed_data = []
        
        try:
            for row in raw_table_data:
                if not row or len(row) == 0:
                    continue
                
                # Первая ячейка - дата
                date_str = row[0] if len(row) > 0 else ""
                
                # Пропускаем строки Period Average, Period High, Period Low
                if date_str.startswith('Period'):
                    continue
                
                # Конвертируем дату в формат DD.MM.YYYY
                formatted_date = self.convert_date_format(date_str)
                
                # Остальные ячейки - курсы валют
                rates = []
                
                # Берем столько курсов, сколько валют
                for i in range(1, min(len(row), num_currencies + 1)):
                    rate = row[i]
                    # Очищаем от лишних символов
                    rate = rate.replace(',', '').strip()
                    rates.append(rate)
                
                # Если курсов меньше чем валют, дополняем пустыми
                while len(rates) < num_currencies:
                    rates.append('')
                
                # Формируем строку: дата + курсы
                parsed_row = [formatted_date] + rates
                parsed_data.append(parsed_row)
            
            print(f"  ✅ Распарсено строк: {len(parsed_data)}")
            if parsed_data and len(parsed_data) > 0:
                print(f"  📊 Пример строки: {parsed_data[0]}")
            
            return parsed_data
            
        except Exception as e:
            print(f"❌ Ошибка парсинга: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def convert_date_format(self, date_str):
        """
        Конвертирует дату из 'Dec 17, 2025' в '17.12.2025'
        
        Args:
            date_str (str): Дата в формате 'Dec 17, 2025'
            
        Returns:
            str: Дата в формате 'DD.MM.YYYY'
        """
        try:
            # Словарь месяцев
            months = {
                'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
                'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
                'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
            }
            
            # Разбиваем строку: 'Dec 17, 2025' -> ['Dec', '17,', '2025']
            parts = date_str.split()
            
            if len(parts) >= 3:
                month_str = parts[0]  # 'Dec'
                day_str = parts[1].replace(',', '')  # '17'
                year_str = parts[2]  # '2025'
                
                # Получаем номер месяца
                month_num = months.get(month_str, '??')
                
                # Форматируем день с ведущим нулем если нужно
                day = day_str.zfill(2)
                
                # Возвращаем в формате DD.MM.YYYY
                return f"{day}.{month_num}.{year_str}"
            else:
                return date_str  # Если не смогли распарсить, возвращаем как есть
                
        except Exception as e:
            print(f"⚠️ Ошибка конвертации даты '{date_str}': {e}")
            return date_str
    
    def save_full_table_csv(self, table_data, filename="oanda_full_table.csv"):
        """Сохранить полные данные таблицы в CSV"""
        if not table_data:
            return False
            
        try:
            with open(filename, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Extracted At', datetime.now().isoformat()])
                writer.writerow([])  # Пустая строка
                
                # Записываем все строки
                for row in table_data:
                    writer.writerow(row)
            
            print(f"💾 Полная таблица сохранена: {filename}")
            return True
        except Exception as e:
            print(f"❌ Ошибка сохранения полной таблицы: {e}")
            return False
    
    def save_to_excel(self, table_data, currency_pair, filename="oanda_rates.xlsx"):
        """
        Сохранить данные в Excel с указанием валютной пары
        
        Args:
            table_data (list): Данные таблицы
            currency_pair (str): Валютная пара (например USD/EUR)
            filename (str): Имя файла
        """
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            import os
            
            # Проверяем, существует ли файл
            if os.path.exists(filename):
                wb = load_workbook(filename)
            else:
                wb = Workbook()
                # Удаляем стандартный лист
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            
            # Создаем лист для валютной пары или используем существующий
            sheet_name = currency_pair.replace('/', '_')
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.delete_rows(1, ws.max_row)  # Очищаем лист
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Заголовок
            ws['A1'] = f'Валютная пара: {currency_pair}'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f'Дата извлечения: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws['A2'].font = Font(italic=True)
            
            # Заголовки таблицы (если есть)
            if table_data:
                start_row = 4
                
                # Добавляем заголовки колонок
                headers = ['Дата', 'Курс', 'Bid', 'Ask', 'High', 'Low']  # Примерные заголовки
                for col_idx, header in enumerate(headers[:len(table_data[0])], 1):
                    cell = ws.cell(row=start_row, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Данные
                for row_idx, row_data in enumerate(table_data, start_row + 1):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Автоширина колонок
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем
            wb.save(filename)
            print(f"📊 Excel сохранен: {filename} (лист: {sheet_name})")
            return True
            
        except Exception as e:
            print(f"❌ Ошибка сохранения Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def save_multi_currency_excel(self, table_data, currency_pairs, filename="oanda_rates.xlsx", sheet_name="USD_Base", log_file="update_log.txt"):
        """
        Умное сохранение с дописыванием новых дат и проверкой старых
        
        Args:
            table_data (list): Распарсенные данные [[дата, курс1, курс2, ...], ...]
            currency_pairs (list): Список валютных пар
            filename (str): Имя файла
            sheet_name (str): Название листа
            log_file (str): Файл для логов
        """
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            import os
            from datetime import datetime
            
            log_messages = []
            log_messages.append(f"\n{'='*80}")
            log_messages.append(f"📋 Обновление листа: {sheet_name}")
            log_messages.append(f"⏰ Время: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            log_messages.append(f"{'='*80}\n")
            
            # Проверяем существует ли файл
            file_exists = os.path.exists(filename)
            
            if file_exists:
                wb = load_workbook(filename)
                log_messages.append(f"📂 Файл найден: {filename}")
            else:
                wb = Workbook()
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
                log_messages.append(f"📄 Создан новый файл: {filename}")
            
            # Создаем или получаем лист
            sheet_exists = sheet_name in wb.sheetnames
            
            if sheet_exists:
                ws = wb[sheet_name]
                log_messages.append(f"📊 Лист '{sheet_name}' существует - обновляем")
                
                # Читаем существующие данные
                existing_data = {}
                start_data_row = 6  # Данные начинаются с 6-й строки
                
                for row_idx in range(start_data_row, ws.max_row + 1):
                    date_cell = ws.cell(row=row_idx, column=1).value
                    if date_cell:
                        # Сохраняем всю строку по дате
                        row_data = []
                        for col_idx in range(1, len(currency_pairs) + 2):
                            cell_value = ws.cell(row=row_idx, column=col_idx).value
                            row_data.append(cell_value)
                        existing_data[str(date_cell)] = row_data
                
                log_messages.append(f"📋 Найдено существующих строк: {len(existing_data)}")
                
            else:
                ws = wb.create_sheet(sheet_name)
                existing_data = {}
                log_messages.append(f"📄 Создан новый лист: {sheet_name}")
            
            # Очищаем лист
            ws.delete_rows(1, ws.max_row)
            
            # Создаем заголовок
            base_currency = currency_pairs[0].split('/')[0] if currency_pairs else "?"
            ws['A1'] = f'Курсы валют Oanda (База: {base_currency})'
            ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws['A2'] = f'Последнее обновление: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}'
            ws['A2'].font = Font(italic=True, size=11)
            
            ws['A3'] = f'Валютные пары: {", ".join(currency_pairs)}'
            ws['A3'].font = Font(italic=True, size=11)
            
            # Объединяем заголовок
            merge_to_col = min(len(currency_pairs) + 1, 26)
            ws.merge_cells(f'A1:{chr(64 + merge_to_col)}1')
            
            # Заголовки таблицы
            start_row = 5
            ws.cell(row=start_row, column=1, value='Дата')
            ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
            ws.cell(row=start_row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')
            
            for col_idx, pair in enumerate(currency_pairs, 2):
                cell = ws.cell(row=start_row, column=col_idx, value=pair)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Обрабатываем данные
            new_dates = 0
            updated_dates = 0
            mismatches = 0
            
            # Сортируем данные по дате (старые сверху, новые внизу)
            sorted_data = sorted(table_data, key=lambda x: self.parse_date_for_sort(x[0]))
            
            for row_idx, row_data in enumerate(sorted_data, start_row + 1):
                date_str = row_data[0]
                
                # Проверяем есть ли эта дата в старых данных
                if date_str in existing_data:
                    old_row = existing_data[date_str]
                    
                    # Проверяем совпадают ли курсы
                    matches = True
                    for i in range(1, len(row_data)):
                        if i < len(old_row):
                            try:
                                old_val = float(str(old_row[i]))
                                new_val = float(str(row_data[i]))
                                if abs(old_val - new_val) > 0.0001:  # Допуск на погрешность
                                    matches = False
                                    log_messages.append(f"⚠️ Несовпадение {date_str}: {currency_pairs[i-1] if i-1 < len(currency_pairs) else '?'} старое={old_val:.4f}, новое={new_val:.4f}")
                                    mismatches += 1
                            except:
                                pass
                    
                    if matches:
                        updated_dates += 1
                else:
                    new_dates += 1
                    log_messages.append(f"✅ Новая дата: {date_str}")
                
                # Записываем строку
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    if col_idx == 1:  # Дата
                        cell.alignment = Alignment(horizontal='left')
                    else:  # Курсы
                        cell.alignment = Alignment(horizontal='right')
                        try:
                            float_val = float(str(value).replace(',', ''))
                            cell.value = float_val
                            cell.number_format = '0.0000'
                        except:
                            pass
            
            # Авторазмер колонок
            for column in ws.columns:
                max_length = 0
                column_letter = None
                
                for cell in column:
                    try:
                        if hasattr(cell, 'column_letter'):
                            if column_letter is None:
                                column_letter = cell.column_letter
                            
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                if column_letter:
                    adjusted_width = min(max_length + 3, 50)
                    ws.column_dimensions[column_letter].width = max(adjusted_width, 15)
            
            # Сохраняем
            wb.save(filename)
            
            # Логи
            log_messages.append(f"\n📊 Статистика обновления:")
            log_messages.append(f"  ✅ Новых дат: {new_dates}")
            log_messages.append(f"  🔄 Обновлено совпадающих: {updated_dates}")
            log_messages.append(f"  ⚠️ Несовпадений: {mismatches}")
            log_messages.append(f"  📋 Всего строк: {len(sorted_data)}")
            log_messages.append(f"\n✅ Excel обновлен: {filename} (лист: {sheet_name})\n")
            
            print("\n".join(log_messages))
            
            # Записываем в лог-файл
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write("\n".join(log_messages))
            
            # Возвращаем статистику
            return {
                'new_dates': new_dates,
                'updated_dates': updated_dates,
                'mismatches': mismatches,
                'total_rows': len(sorted_data)
            }
            
        except Exception as e:
            error_msg = f"❌ Ошибка сохранения Excel: {e}"
            print(error_msg)
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(f"\n{error_msg}\n")
            import traceback
            traceback.print_exc()
            return None
    
    def parse_date_for_sort(self, date_str):
        """
        Конвертирует дату DD.MM.YYYY в число для сортировки
        
        Args:
            date_str (str): Дата '17.12.2025'
            
        Returns:
            int: Число для сортировки (например 20251217)
        """
        try:
            parts = date_str.split('.')
            if len(parts) == 3:
                day, month, year = parts
                # Возвращаем YYYYMMDD как число
                return int(f"{year}{month.zfill(2)}{day.zfill(2)}")
            return 0
        except:
            return 0
        """
        Сохранить данные всех валютных пар в Excel файл (на определенный лист)
        Все валюты в столбцах, даты в строках
        
        Args:
            table_data (list): Распарсенные данные [[дата, курс1, курс2, ...], ...]
            currency_pairs (list): Список валютных пар (например ['USD/EUR', 'USD/AED'])
            filename (str): Имя файла
            sheet_name (str): Название листа (например 'USD_Base', 'EUR_Base')
        """
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            import os
            
            # Проверяем существует ли файл
            if os.path.exists(filename):
                wb = load_workbook(filename)
            else:
                wb = Workbook()
                # Удаляем стандартный лист
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            
            # Создаем новый лист или используем существующий
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Очищаем лист
                ws.delete_rows(1, ws.max_row)
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Заголовок
            base_currency = currency_pairs[0].split('/')[0] if currency_pairs else "?"
            ws['A1'] = f'Курсы валют Oanda (База: {base_currency})'
            ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws['A2'] = f'Дата извлечения: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws['A2'].font = Font(italic=True, size=11)
            
            ws['A3'] = f'Валютные пары: {", ".join(currency_pairs)}'
            ws['A3'].font = Font(italic=True, size=11)
            
            # Объединяем ячейки заголовка
            merge_to_col = min(len(currency_pairs) + 1, 26)  # Не больше Z
            ws.merge_cells(f'A1:{chr(64 + merge_to_col)}1')
            
            # Пустая строка
            start_row = 5
            
            # Заголовки таблицы
            # Первая колонка - Дата
            ws.cell(row=start_row, column=1, value='Дата')
            ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
            ws.cell(row=start_row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')
            
            # Заголовки валютных пар
            for col_idx, pair in enumerate(currency_pairs, 2):
                cell = ws.cell(row=start_row, column=col_idx, value=pair)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Данные таблицы
            if table_data:
                for row_idx, row_data in enumerate(table_data, start_row + 1):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Форматирование
                        if col_idx == 1:  # Дата
                            cell.alignment = Alignment(horizontal='left')
                        else:  # Курсы
                            cell.alignment = Alignment(horizontal='right')
                            # Если это число, форматируем
                            try:
                                float_val = float(str(value).replace(',', ''))
                                cell.value = float_val
                                cell.number_format = '0.0000'
                            except:
                                pass
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = None
                
                for cell in column:
                    try:
                        # Пропускаем объединенные ячейки
                        if hasattr(cell, 'column_letter'):
                            if column_letter is None:
                                column_letter = cell.column_letter
                            
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                if column_letter:
                    adjusted_width = min(max_length + 3, 50)
                    ws.column_dimensions[column_letter].width = max(adjusted_width, 15)
            
            # Сохраняем
            wb.save(filename)
            print(f"✅ Excel обновлен: {filename} (лист: {sheet_name})")
            print(f"   📊 Валютных пар: {len(currency_pairs)}")
            print(f"   📋 Строк данных: {len(table_data)}")
            return True
            
        except Exception as e:
            print(f"❌ Ошибка сохранения Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def get_all_rates(self):
        """Получить все курсы"""
        rates = {}
        
        # Основной метод - из таблицы
        table_rates = self.extract_rates_from_table()
        rates.update(table_rates)
        
        # Если ничего не нашли, пробуем альтернативный метод
        if not rates:
            print("⚠️ Таблица пуста, пробую альтернативный метод...")
            alt_rates = self.extract_rates_alternative()
            rates.update(alt_rates)
        
        return rates
    
    def save_screenshot(self, filename="oanda_screenshot.png"):
        """Сохранение скриншота"""
        try:
            self.driver.save_screenshot(filename)
            print(f"📸 Скриншот: {filename}")
            return True
        except Exception as e:
            print(f"❌ Ошибка скриншота: {e}")
            return False
    
    def save_html(self, filename="oanda_page.html"):
        """Сохранение HTML"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(self.driver.page_source)
            print(f"💾 HTML: {filename}")
            return True
        except Exception as e:
            print(f"❌ Ошибка сохранения HTML: {e}")
            return False
    
    def save_json(self, rates, filename="currency_rates.json"):
        """Сохранение в JSON"""
        try:
            data = {
                "timestamp": datetime.now().isoformat(),
                "source": self.url,
                "total_pairs": len(rates),
                "rates": rates
            }
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            print(f"💾 JSON: {filename}")
            return True
        except Exception as e:
            print(f"❌ Ошибка сохранения JSON: {e}")
            return False
    
    def save_csv(self, rates, filename="currency_rates.csv"):
        """Сохранение в CSV"""
        try:
            with open(filename, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Timestamp', 'Currency Pair', 'Rate'])
                
                timestamp = datetime.now().isoformat()
                for pair, rate in sorted(rates.items()):
                    writer.writerow([timestamp, pair, rate])
            
            print(f"💾 CSV: {filename}")
            return True
        except Exception as e:
            print(f"❌ Ошибка сохранения CSV: {e}")
            return False
    
    def print_rates(self, rates):
        """Красивый вывод курсов"""
        if not rates:
            return
            
        print("\n" + "="*60)
        print("💱 КУРСЫ ВАЛЮТ OANDA:")
        print("="*60)
        for pair, rate in sorted(rates.items()):
            print(f"  {pair:15} → {rate:>12}")
        print("="*60 + "\n")
    
    def close(self):
        """Закрытие браузера"""
        if self.driver:
            self.driver.quit()
            print("🔒 Браузер закрыт")



def upload_to_github(excel_file, log_file, script_file):
    """Загрузка файлов на GitHub"""
    
    import os
    import subprocess
    import shutil
    
    GITHUB_TOKEN = "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"
    REPO_URL = "https://github.com/KMS-wq/123"
    REPO_NAME = "123"
    
    files_to_upload = [excel_file, log_file, script_file]
    
    try:
        
        # Проверяем файлы
        print("🔍 Проверяю файлы...")
        for file in files_to_upload:
            if os.path.exists(file):
                size = os.path.getsize(file) / 1024
                print(f"  ✅ {file} ({size:.1f} KB)")
            else:
                print(f"  ❌ {file} - не найден!")
                return False
        
        # Настройка git
        print("\n🔧 Настраиваю git...")
        subprocess.run(["git", "config", "--global", "user.email", "bot@oanda.local"], 
                      capture_output=True, check=True)
        subprocess.run(["git", "config", "--global", "user.name", "Oanda Bot"], 
                      capture_output=True, check=True)
        
        # Клонируем репозиторий
        print(f"\n📥 Клонирую {REPO_NAME}...")
        repo_url_with_token = REPO_URL.replace("https://", f"https://{GITHUB_TOKEN}@")
        
        if os.path.exists(REPO_NAME):
            shutil.rmtree(REPO_NAME)
        
        result = subprocess.run(
            ["git", "clone", repo_url_with_token],
            capture_output=True,
            text=True
        )
        
        if result.returncode != 0:
            print(f"❌ Ошибка: {result.stderr}")
            return False
        
        print("✅ Репозиторий клонирован")
        
        # Копируем файлы
        print("\n📋 Копирую файлы...")
        for file in files_to_upload:
            dest = os.path.join(REPO_NAME, file)
            shutil.copy2(file, dest)
            print(f"  ✅ {file}")
        
        # Создаем README
        readme_path = os.path.join(REPO_NAME, "README.md")
        if not os.path.exists(readme_path):
            with open(readme_path, 'w', encoding='utf-8') as f:
                f.write(f"""# 📊 Oanda Currency Rates Data

Автоматический сбор курсов валют с Oanda.

## 📁 Файлы:

- **{excel_file}** - данные по 5 базовым валютам
- **{log_file}** - лог обновлений
- **{script_file}** - скрипт сбора данных

## 📊 Данные:

- **180 дней** истории
- **5 баз**: USD, EUR, AED, GBP, HKD  
- **10 валют** к каждой базе
- **Формат даты**: DD.MM.YYYY

## 🔄 Обновление:

{datetime.now().strftime('%d.%m.%Y %H:%M:%S')}
""")
            print("  ✅ README.md")
        
        # Git операции
        os.chdir(REPO_NAME)
        
        print("\n➕ Git add...")
        subprocess.run(["git", "add", "."], check=True)
        
        print("💾 Git commit...")
        commit_msg = f"Update {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        subprocess.run(["git", "commit", "-m", commit_msg], check=True)
        
        print("📤 Git push...")
        result = subprocess.run(["git", "push", "origin", "main"], 
                               capture_output=True, text=True)
        
        if result.returncode != 0:
            result = subprocess.run(["git", "push", "origin", "master"],
                                   capture_output=True, text=True)
        
        os.chdir("..")
        
        if result.returncode == 0:
            print(f"\n{'='*80}")
            print("✅ ЗАГРУЖЕНО НА GITHUB!")
            print(f"{'='*80}")
            print(f"🔗 {REPO_URL}")
            print(f"{'='*80}\n")
            return True
        else:
            print(f"❌ Ошибка push: {result.stderr}")
            return False
            
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Главная функция с циклом по базовым валютам"""
    print("\n" + "="*80)
    print("🌍 OANDA MULTI-BASE CURRENCY SCRAPER")
    print("="*80)
    print(f"⏰ Время: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*80 + "\n")
    
    # Конфигурация базовых валют
    base_currencies_config = [
        ("USD", None, "USD_Base"),      # USD - по умолчанию
        ("EUR", 2, "EUR_Base"),         # EUR - li[2]
        ("AED", 10, "AED_Base"),        # AED - li[10]
        ("GBP", 83, "GBP_Base"),        # GBP - li[83]
        ("HKD", 94, "HKD_Base"),        # HKD - li[94]
    ]
    
    # Конфигурация валют в столбцах
    currencies_config = [
        (2, "AED", 10, None),
        (3, "CHF", 48, None),
        (4, "GBP", 83, None),
        (5, "JPY", 111, None),
    ]
    
    # Плюсы для дополнительных столбцов
    plus_buttons_config = [6, 7, 8, 9, 10]
    
    # Валюты для плюсов
    extra_currencies_config = [
        (6, "INR", 103),
        (7, "RUB", 179),
        (8, "KZT", 121),
        (9, "HKD", 94),
        (10, "USD", 3),
    ]
    
    excel_filename = "oanda_all_bases_currencies.xlsx"
    log_filename = "update_log.txt"
    scraper = OandaRateScraperV2(headless=False, wait_time=20)
    success_count = 0
    
    # Статистика для итогового лога
    total_stats = {
        'total_new': 0,
        'total_updated': 0,
        'total_mismatches': 0,
        'total_rows': 0,
        'sheets_processed': []
    }
    
    try:
        # Записываем заголовок лога
        with open(log_filename, 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write("🌍 OANDA CURRENCY RATES UPDATE LOG\n")
            f.write("="*80 + "\n")
            f.write(f"⏰ Время запуска: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n")
            f.write(f"📁 Файл: {excel_filename}\n")
            f.write("="*80 + "\n\n")
            f.write("📊 ИТОГОВАЯ СТАТИСТИКА (будет заполнена после обработки)\n")
            f.write("="*80 + "\n\n")
        
        # Запуск браузера
        if not scraper.start():
            sys.exit(1)
        
        if not scraper.load_page():
            sys.exit(1)
        
        scraper.accept_cookies()
        scraper.select_time_period(180)
        scraper.switch_to_table_view()
        
        # ЦИКЛ ПО БАЗОВЫМ ВАЛЮТАМ
        for base_idx, (base_curr, base_li, sheet_name) in enumerate(base_currencies_config, 1):
            print(f"\n{'='*80}")
            print(f"🌍 БАЗА {base_idx}/{len(base_currencies_config)}: {base_curr}")
            print(f"{'='*80}\n")
            
            # Меняем базовую валюту (кроме USD)
            if base_li is not None:
                if not scraper.change_base_currency(base_curr, base_li):
                    print(f"⚠️ Не удалось сменить на {base_curr}")
                    continue
                print(f"✅ Базовая валюта: {base_curr}\n")
                time.sleep(3)
            
            # Формируем список пар для текущей базы
            current_pairs = [f"{base_curr}/EUR"]  # EUR всегда вторая по умолчанию
            
            # ТОЛЬКО ДЛЯ ПЕРВОЙ БАЗЫ (USD) - добавляем столбцы
            if base_idx == 1:
                # Добавляем 4 валюты
                for col_num, curr_code, li_idx, _ in currencies_config:
                    print(f"➕ Добавляю {base_curr}/{curr_code}...")
                    if scraper.add_currency_column(col_num, curr_code, li_idx):
                        print(f"✅ {base_curr}/{curr_code}")
                        current_pairs.append(f"{base_curr}/{curr_code}")
                    time.sleep(1.5)
                
                # Добавляем 5 плюсов
                print(f"\n➕ Добавляю 5 столбцов...")
                scraper.add_multiple_plus_buttons(plus_buttons_config)
                
                # Выбираем валюты в новых столбцах
                for div_num, curr_code, li_idx in extra_currencies_config:
                    print(f"💱 Выбираю {base_curr}/{curr_code}...")
                    if scraper.select_currency_in_new_column(div_num, curr_code, li_idx):
                        print(f"✅ {base_curr}/{curr_code}")
                        current_pairs.append(f"{base_curr}/{curr_code}")
                    time.sleep(1.5)
            else:
                # Для остальных баз - просто добавляем названия пар
                for _, curr_code, _, _ in currencies_config:
                    current_pairs.append(f"{base_curr}/{curr_code}")
                for _, curr_code, _ in extra_currencies_config:
                    current_pairs.append(f"{base_curr}/{curr_code}")
            
            # Скриншот
            scraper.save_screenshot(f"oanda_{base_curr}_base.png")
            
            # Извлекаем данные
            print(f"\n📊 Извлекаю данные для {base_curr}...")
            raw_data = scraper.extract_full_table_data()
            
            if raw_data:
                parsed_data = scraper.parse_table_data(raw_data, len(current_pairs))
                
                if parsed_data:
                    # Сохраняем на отдельный лист
                    sheet_stats = scraper.save_multi_currency_excel(
                        parsed_data,
                        current_pairs,
                        excel_filename,
                        sheet_name,
                        log_filename
                    )
                    
                    # Собираем статистику
                    if sheet_stats:
                        total_stats['sheets_processed'].append(sheet_name)
                        total_stats['total_new'] += sheet_stats.get('new_dates', 0)
                        total_stats['total_updated'] += sheet_stats.get('updated_dates', 0)
                        total_stats['total_mismatches'] += sheet_stats.get('mismatches', 0)
                        total_stats['total_rows'] += sheet_stats.get('total_rows', 0)
                    
                    success_count += 1
                    print(f"✅ Лист '{sheet_name}' сохранен!\n")
                else:
                    print(f"⚠️ Не удалось распарсить данные для {base_curr}\n")
            else:
                print(f"⚠️ Не удалось извлечь данные для {base_curr}\n")
            
            # Пауза перед следующей базой
            if base_idx < len(base_currencies_config):
                time.sleep(2)
        
        # Записываем итоговую статистику в начало файла
        with open(log_filename, 'r', encoding='utf-8') as f:
            log_content = f.read()
        
        summary = []
        summary.append("="*80 + "\n")
        summary.append("📊 ИТОГОВАЯ СТАТИСТИКА ПО ВСЕМ ЛИСТАМ\n")
        summary.append("="*80 + "\n")
        summary.append(f"✅ Обработано листов: {success_count}/{len(base_currencies_config)}\n")
        summary.append(f"📋 Листы: {', '.join(total_stats['sheets_processed'])}\n")
        summary.append(f"\n💾 Общие цифры по всем листам:\n")
        summary.append(f"  ✅ Всего новых дат: {total_stats['total_new']}\n")
        summary.append(f"  🔄 Всего обновлено: {total_stats['total_updated']}\n")
        summary.append(f"  ⚠️ Всего несовпадений: {total_stats['total_mismatches']}\n")
        summary.append(f"  📊 Всего строк данных: {total_stats['total_rows']}\n")
        summary.append("="*80 + "\n\n")
        
        # Вставляем итоговую статистику после заголовка
        parts = log_content.split("📊 ИТОГОВАЯ СТАТИСТИКА (будет заполнена после обработки)\n" + "="*80 + "\n\n", 1)
        if len(parts) == 2:
            new_content = parts[0] + "".join(summary) + parts[1]
            with open(log_filename, 'w', encoding='utf-8') as f:
                f.write(new_content)
        
        # ИТОГИ
        print(f"\n{'='*80}")
        print(f"✅ ЗАВЕРШЕНО!")
        print(f"{'='*80}")
        print(f"📊 Обработано баз: {success_count}/{len(base_currencies_config)}")
        print(f"📁 Файл Excel: {excel_filename}")
        print(f"📄 Файл лога: {log_filename}")
        print(f"{'='*80}\n")
        
        success = success_count > 0
        time.sleep(3)
        
    except KeyboardInterrupt:
        print("\n⚠️ Прервано пользователем")
        success = False
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        success = False
    finally:
        scraper.close()
        
        print("\n" + "="*80)
        if success:
            print("✅ Работа завершена успешно")
        else:
            print("⚠️ Работа завершена с ошибками")
        print("="*80 + "\n")
    
    # Автоматическая загрузка на GitHub
    if success:
        print(f"\n{'='*80}")
        print("📤 ЗАГРУЗКА НА GITHUB")
        print(f"{'='*80}\n")
        
        upload_success = upload_to_github(
            excel_filename,
            log_filename,
            "oanda_multi_base_final.py"
        )
        
        if upload_success:
            print("\n🎉 Данные успешно загружены на GitHub!")
        else:
            print("\n⚠️ Загрузка на GitHub не удалась")


if __name__ == "__main__":
    main()


